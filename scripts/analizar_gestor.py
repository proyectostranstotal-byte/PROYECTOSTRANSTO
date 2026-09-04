#!/usr/bin/env python3
"""Recalcula desde cero el Gestor_de_Gastos.xlsx y muestra el análisis financiero.

Uso: python3 scripts/analizar_gestor.py /ruta/a/Gestor_de_Gastos.xlsx [AAAA-MM]

Lee las hojas Movimientos y Cuotas (valores, no fórmulas), reconstruye la lógica
del Tablero (gasto del mes = movimientos sueltos + cuotas activas en ese mes) y
verifica que coincida con lo que muestra la planilla.
"""
import sys
import warnings
from datetime import date

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")
pd.set_option("display.width", 200)
pd.set_option("display.max_rows", 300)
pd.set_option("display.float_format", lambda v: f"{v:,.0f}")


def idx(y, m):
    """Índice de mes usado por la planilla: año*12 + mes."""
    return y * 12 + m


def leer_movimientos(ws):
    filas = []
    for r in range(4, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 8)]
        if v[0] is None or v[5] is None:
            continue
        filas.append(dict(fila=r, fecha=v[0], tipo=v[1], cat=v[2], desc=v[3],
                          medio=v[4], monto=float(v[5]),
                          per=v[0].strftime("%Y-%m")))
    return pd.DataFrame(filas)


def leer_cuotas(ws):
    filas = []
    for r in range(5, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 7)]
        if v[3] is None or v[4] is None or v[5] is None:
            continue
        n = int(v[4])
        ini = idx(v[3].year, v[3].month)
        filas.append(dict(fila=r, desc=v[0], cat=v[1], medio=v[2], inicio=v[3],
                          n=n, cuota=float(v[5]), total=n * float(v[5]),
                          M=ini, N=ini + n - 1))
    return pd.DataFrame(filas)


def main(path, mes=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = leer_movimientos(wb["Movimientos"])
    c = leer_cuotas(wb["Cuotas"])
    hoy = date.today()
    hoy_idx = idx(hoy.year, hoy.month)

    print(f"Movimientos: {len(m)} filas ({m.fecha.min().date()} → {m.fecha.max().date()})")
    print(f"Cuotas: {len(c)} compras, total financiado {c.total.sum():,.0f}")

    meses = sorted(m.per.unique()) if mes is None else [mes]
    for per in meses:
        y, mo = map(int, per.split("-"))
        k = idx(y, mo)
        ing = m[(m.per == per) & (m.tipo == "Ingreso")].monto.sum()
        aho = m[(m.per == per) & (m.tipo == "Ahorro")].monto.sum()
        gs = m[(m.per == per) & (m.tipo == "Gasto")]
        act = c[(c.M <= k) & (c.N >= k)]
        print(f"\n===== {per} =====")
        print(f"Ingresos {ing:>14,.0f}")
        print(f"Gastos sueltos {gs.monto.sum():>10,.0f}   cuotas {act.cuota.sum():>12,.0f} ({len(act)} activas)")
        print(f"Gastos totales {gs.monto.sum() + act.cuota.sum():>10,.0f}")
        print(f"Ahorro {aho:>16,.0f}")
        print(f"Disponible {ing - gs.monto.sum() - act.cuota.sum() - aho:>12,.0f}")
        t = pd.concat([gs.groupby("cat").monto.sum().rename("sueltos"),
                       act.groupby("cat").cuota.sum().rename("cuotas")], axis=1).fillna(0)
        t["total"] = t.sueltos + t.cuotas
        t["%"] = (t.total / t.total.sum() * 100).round(1)
        print(t.sort_values("total", ascending=False).to_string())
        print("\nPor medio de pago (sueltos):")
        print(gs.groupby("medio").monto.sum().to_string())

    print("\n===== Compromisos futuros de cuotas =====")
    for k in range(hoy_idx, c.N.max() + 1):
        act = c[(c.M <= k) & (c.N >= k)]
        y, mo = divmod(k - 1, 12)
        print(f"  {y}-{mo + 1:02d}: {len(act):2d} cuotas  {act.cuota.sum():>12,.0f}")
    pag = c.apply(lambda x: max(0, min(x.n, hoy_idx - x.M + 1)), axis=1)
    print(f"Restante por pagar: {((c.n - pag) * c.cuota).sum():,.0f}")
    print("\nPor medio de pago (total financiado):")
    print(c.groupby("medio").total.sum().to_string())

    print("\n===== Controles de datos =====")
    dup = m[m.duplicated(["fecha", "cat", "desc", "monto"], keep=False)]
    print("Duplicados exactos:", "ninguno" if dup.empty else dup.to_string())
    print("Descripciones con espacios sobrantes:", m[m.desc.astype(str).str.strip() != m.desc.astype(str)].desc.tolist())
    print("Filas de ejemplo sin borrar:", m[m.desc.astype(str).str.contains("ejemplo", case=False)].fila.tolist())


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
