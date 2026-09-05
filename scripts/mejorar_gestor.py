#!/usr/bin/env python3
"""Mejora el Gestor_de_Gastos.xlsx: Billetera acumulada, Pagos por tarjeta y cuotas marcadas.

Uso: python3 scripts/mejorar_gestor.py entrada.xlsx salida.xlsx

Cambios que aplica sobre la planilla original (conserva todo lo demás):
  * Nueva hoja "Pagos": por mes, cuánto pagar de Crédito / MercadoPago / Otros
    y una celda amarilla "Pagado el" por tarjeta para marcar la fecha de pago.
  * Nueva hoja "Billetera": saldo inicial + ingresos - gastos - ahorro - tarjetas
    pagadas + fondo Aston Birra = plata que deberías tener hoy; pendientes de
    pago, disponible real, control de caja y sobrante acumulado mes a mes.
  * Cuotas: "Pagadas", "Faltan" y "Estado" se calculan con las marcas de Pagos
    (ya no con la fecha de hoy). Se agregan "Restante ($)" y "Próxima a pagar".
  * Tablero: arregla la fórmula faltante de Supermercado (C18), el TOTAL que
    mezclaba ingresos y gastos, y muestra Crédito / MercadoPago del mes y la
    Billetera.
  * Movimientos: borra la fila de ejemplo y corrige los desplegables de Tipo y
    Medio de pago para que lean Configuración completa.
  * Instrucciones: agrega la explicación del nuevo circuito.
"""
import sys
import warnings
from copy import copy
from datetime import date

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---- paleta de la planilla original -------------------------------------
AZUL_TITULO = "FF1F3A5F"
AZUL_HEADER = "FF2E5C8A"
VERDE = "FF2E7D32"
ROJO = "FFC0392B"
GRIS = "FF7F8C8D"
CELESTE = "FFDCE6F1"
AMARILLO = "FFFFF2CC"
AZUL_INPUT = "FF0000FF"
BORDE = "FFBFC9D4"
MONEDA = '"$ "#,##0;[Red]"-$ "#,##0'
FECHA = "d/m/yyyy"

thin = Side(style="thin", color=BORDE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(rgb):
    return PatternFill(fill_type="solid", start_color=rgb, end_color=rgb)


def style(cell, *, bold=False, size=10, color="FF000000", bg=None, fmt=None,
          h=None, v="center", wrap=False, border=True, italic=False):
    cell.font = Font(name="Arial", size=size, bold=bold, color=color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        cell.border = BORDER


def title(ws, ref, text, size=16):
    ws[ref] = text
    style(ws[ref], bold=True, size=size, color="FFFFFFFF", bg=AZUL_TITULO, h="left", border=False)


def note(ws, ref, text):
    ws[ref] = text
    style(ws[ref], color=GRIS, size=9, italic=True, h="left", wrap=True, border=False)


def header(cell, text, bg=AZUL_HEADER):
    cell.value = text
    style(cell, bold=True, color="FFFFFFFF", bg=bg, h="center", wrap=True)


def money(cell, formula, bold=False, size=10, color="FF000000", bg=None):
    cell.value = formula
    style(cell, fmt=MONEDA, bold=bold, size=size, color=color, bg=bg)


def label(cell, text, bold=False, bg=None, color="FF000000", indent=False):
    cell.value = text
    style(cell, bold=bold, bg=bg, color=color, h="left")
    if indent:
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)


def entrada(cell, value=None, fmt=None):
    """Celda amarilla que completa el usuario (convención de la planilla)."""
    cell.value = value
    style(cell, color=AZUL_INPUT, bg=AMARILLO, fmt=fmt)


def idx(y, m):
    return y * 12 + m


# (año, mes), columna de Pagos, valor. E/I/L = "Pagado el"; F/J/M = "Monto real $"; G = "USD pagados".
PAGOS_REALES = [
    ((2026, 8), "F", 452253),                    # Visa cierre 27/07: pago del 06/08 (436.451,55 + USD 10,43 al 1.515)
    ((2026, 8), "J", 202017.61),                 # Mercado Pago cierre 27/07: débito automático del 06/08
    ((2026, 9), "E", "✔"),                       # Visa cierre 27/08: pagada
    ((2026, 9), "F", 1150000),                   # pesos pagados (resumen 1.174.457,23 menos percepción 30 % = 1.150.407,34)
    ((2026, 9), "G", 52.95),                     # USD comprados con $ 80.000 (incluye los 20 de Anthropic, que se reintegran)
    ((2026, 9), "I", date(2026, 9, 2)),          # Mercado Pago cierre 27/08: débito automático del 02/09
    ((2026, 9), "J", 273984.72),
]

# Correcciones de filas de Cuotas confirmadas contra los resúmenes: (descripción, columna, valor nuevo)
AJUSTES_CUOTAS = [
    ("Omega 3", "C", "MercadoPago"), ("Omega 3", "F", 26412.5),          # MP 06/08 ML 52.825 en 2, mitad cada uno con su novia
    ("Regalo benja", "C", "MercadoPago"), ("Regalo benja", "F", 17449.12),  # MP 02/08 Distrimicabeb 34.898,23 en 2, mitad c/u
    # importes exactos del resumen Visa cierre 27/08/2026
    ("Notebook", "F", 116666.58), ("Regalo fiama", "F", 50382.50), ("Garmin forerunner", "F", 68214.73),
    ("Celular", "F", 233093.72), ("Ojotas", "F", 49999), ("Cena", "F", 17163), ("Seguro", "F", 158178),
    ("Super Ebe", "F", 52807.50), ("Adidas Urban", "F", 55833.35), ("Proteina", "F", 112413.43),
]
# "Regalo fiama" aparece dos veces (junio y septiembre): el ajuste anterior aplica a la primera; la de septiembre va aparte
AJUSTES_CUOTAS_POR_FILA = [(31, "F", 43333.20)]   # fila 31 = Regalo fiama 6 cuotas desde 09/2026 (127-MOOV 43.333,20)

# Suscripciones que se pagan en dólares: descripción → USD por mes (columna R de Cuotas; F pasa a calcularse con el dólar)
CUOTAS_USD = {"Youtube Premium": 6.97, "Spotify": 2.22, "Google One": 2.99, "Pulso": 13.99}

# Gastos del resumen Visa 27/08/2026 que faltaban en Cuotas: (descripción, categoría, medio, fecha 1ª cuota, cuotas, $ por cuota, USD por cuota)
CUOTAS_NUEVAS = [
    ("Comisión renovación anual Visa + IVA", "Impuestos / Comisiones", "Crédito", date(2026, 9, 6), 1, 42324.99, None),
    ("Entreno (suplementos)", "Suplementos / Nutrición", "Crédito", date(2026, 9, 6), 1, 53515, None),
    ("MELI 02/08", "Otros gastos", "Crédito", date(2026, 9, 6), 1, 3490, None),
    ("Impuestos s/ suscripciones USD (IVA 21% + IIBB 3%)", "Impuestos / Comisiones", "Crédito", date(2026, 9, 6), 1, 11972.70, None),
    ("Impuestos s/ suscripciones USD (desde octubre, sin Paramount)", "Impuestos / Comisiones", "Crédito", date(2026, 10, 6), 11, 10575, None),
    ("Impuesto sellos Visa", "Impuestos / Comisiones", "Crédito", date(2026, 9, 6), 12, 1244.29, None),
    ("Saldo resumen anterior sin pagar", "Otros gastos", "Crédito", date(2026, 9, 6), 1, 20483.15, None),
    ("Devolución HBO Max", "Otros gastos", "Crédito", date(2026, 9, 6), 1, -23207.80, None),
    ("Paramount+ (dado de baja)", "Suscripciones (streaming, apps)", "Crédito", date(2026, 9, 6), 1, None, 3.79),
    ("Hevy Gym", "Suscripciones (streaming, apps)", "Crédito", date(2026, 9, 6), 12, None, 2.99),
    ("Pan de masa madre y yerba", "Supermercado", "Crédito", date(2026, 10, 6), 1, 25900, None),
    ("Shorts Monkeyforce", "Ropa / Cuidado personal", "Crédito", date(2026, 10, 6), 6, 33900, None),
]

# Movimientos: correcciones y altas informadas por el usuario
MOVIMIENTOS_AJUSTES = [("Sueldo", 1458586, "E", "Débito / Banco")]          # (descripción, monto, columna, valor)
MOVIMIENTOS_NUEVOS = [(date(2026, 9, 4), "Gasto", "Comida / Delivery / Restaurantes", "Cena", "Efectivo", 35500)]
# Billetera: plata que te deben y facturas previstas del mes (editables)
POR_COBRAR = [("Reintegro Anthropic (trabajo)", 30600), ("Fiama (Splitwise: Omega 3 y regalos)", 25000)]
PREVISTOS = [("Luz (Epe) — estimado según agosto", 138226), ("Celular / Internet (Personal) — estimado según agosto", 125000),
             ("Agua — estimado según agosto", 56261)]

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto",
         "Septiembre", "Octubre", "Noviembre", "Diciembre"]


# =========================================================================
TC = "'Configuración'!$B$4"


def main(src, dst):
    wb = openpyxl.load_workbook(src)
    cfg, mov, cuo, tab, ins, ast = (wb["Configuración"], wb["Movimientos"], wb["Cuotas"],
                                    wb["Tablero"], wb["Instrucciones"], wb["Aston Birra"])
    anio = int(cfg["B3"].value)
    cfg["A4"] = "Dólar (para pasar a pesos lo que pagás en USD):"
    cfg["A4"]._style = copy(cfg["A3"]._style)
    entrada(cfg["B4"], 1511, fmt='"$ "#,##0')   # 80.000 / 52,95 USD
    cfg["D4"] = "◄ ponelo al cambio del día que comprás los dólares"
    cfg["D4"]._style = copy(cfg["D3"]._style)
    hoy = date.today()
    hoy_idx = idx(hoy.year, hoy.month)

    # ------------------------------------------------------------ Movimientos
    if mov["D4"].value and "ejemplo" in str(mov["D4"].value).lower():
        mov.delete_rows(4)
        for r in range(4, 605):
            mov[f"G{r}"] = f'=IF($A{r}="","",TEXT($A{r},"YYYY-MM"))'
        for col in "ABCDEFG":
            mov[f"{col}604"]._style = copy(mov[f"{col}603"]._style)
    for desc, monto, col, val in MOVIMIENTOS_AJUSTES:
        for r in range(4, 605):
            if str(mov[f"D{r}"].value or "").strip().lower() == desc.lower() and mov[f"F{r}"].value == monto:
                mov[f"{col}{r}"] = val
                break
    libre = next(r for r in range(4, 605) if mov[f"A{r}"].value is None)
    for fila in MOVIMIENTOS_NUEVOS:
        for col, val in zip("ABCDEF", fila):
            mov[f"{col}{libre}"] = val
        libre += 1
    for dv in mov.data_validations.dataValidation:
        if "E4" in str(dv.sqref):
            dv.formula1 = "'Configuración'!$E$7:$E$13"
        if "B4" in str(dv.sqref):
            dv.formula1 = "'Configuración'!$F$7:$F$10"

    # ------------------------------------------------------------------ Pagos
    # A idx(oculta) | B Mes | C Crédito $ | D Crédito USD | E Pagado el | F Monto real $ | G USD pagados |
    # H MP $ | I Pagado el | J Monto real $ | K Otros $ | L Pagado el | M Monto real $ |
    # N Total $ del mes | O Pendiente $ | P Estado | Q Diferencia $ | R auxiliar (hoy)
    pg = wb.create_sheet("Pagos")
    pg.sheet_view.showGridLines = False
    pg.merge_cells("A1:Q1")
    title(pg, "A1", "💳  Pagos  —  qué pagar cada mes (en pesos y en dólares), cuándo y cuánto pagaste")
    pg.row_dimensions[1].height = 30
    pg.merge_cells("A2:Q2")
    note(pg, "A2", "Cada fila es un mes. 'A pagar' sale solo de la hoja Cuotas: la Visa se separa en pesos y en dólares (los USD "
         "se pasan a pesos con el dólar de Configuración). Cuando pagues, escribí la fecha (o ✔) en 'Pagado el', los pesos que "
         "pagaste en 'Monto real $' y los dólares que compraste en 'USD pagados'. La Billetera descuenta eso, y 'Diferencia' "
         "muestra cuánto quedó sin cargar en Cuotas. Los meses anteriores a la creación de esta hoja ya quedaron marcados con ✔.")
    pg.row_dimensions[2].height = 60
    pg["R1"] = "=YEAR(TODAY())*12+MONTH(TODAY())"   # índice del mes actual (auxiliar)
    pg["R2"] = "hoy (auxiliar)"
    pg.column_dimensions["R"].hidden = True

    heads = ["idx", "Mes",
             "💳 Crédito\na pagar ($)", "💳 Crédito\na pagar (USD)", "Pagado el\n(fecha o ✔)", "Monto real\npagado ($)", "USD\npagados",
             "🟡 MercadoPago\na pagar ($)", "Pagado el\n(fecha o ✔)", "Monto real\npagado ($)",
             "Otros a pagar\n(préstamo…)", "Pagado el\n(fecha o ✔)", "Monto real\npagado ($)",
             "Total del mes ($)\n(USD al dólar config.)", "Pendiente ($)", "Estado", "Diferencia ($)\nreal − calculado"]
    for i, h in enumerate(heads, 1):
        header(pg.cell(4, i), h)
    pg.row_dimensions[4].height = 42
    FIRST, N_MESES = 5, 36
    LAST = FIRST + N_MESES - 1
    USD_FMT = '"USD "#,##0.00'
    y, m = anio - 1, 1
    for r in range(FIRST, LAST + 1):
        pg[f"A{r}"] = idx(y, m)
        pg[f"B{r}"] = f"{MESES[m - 1]} {y}"
        style(pg[f"B{r}"], bold=True, h="left")
        cu = f"(Cuotas!$M$5:$M$102<=$A{r})*(Cuotas!$N$5:$N$102>=$A{r})"
        money(pg[f"C{r}"], f'=SUMPRODUCT({cu}*(Cuotas!$C$5:$C$102="Crédito")*(Cuotas!$R$5:$R$102="")*Cuotas!$F$5:$F$102)')
        pg[f"D{r}"] = f'=SUMPRODUCT({cu}*(Cuotas!$C$5:$C$102="Crédito")*Cuotas!$R$5:$R$102)'
        style(pg[f"D{r}"], fmt=USD_FMT, color=AZUL_TITULO)
        entrada(pg[f"E{r}"], fmt=FECHA)
        entrada(pg[f"F{r}"], fmt=MONEDA)
        entrada(pg[f"G{r}"], fmt=USD_FMT)
        money(pg[f"H{r}"], f'=SUMPRODUCT({cu}*(Cuotas!$C$5:$C$102="MercadoPago")*Cuotas!$F$5:$F$102)')
        entrada(pg[f"I{r}"], fmt=FECHA)
        entrada(pg[f"J{r}"], fmt=MONEDA)
        money(pg[f"K{r}"], f"=SUMPRODUCT({cu}*Cuotas!$F$5:$F$102)-C{r}-D{r}*{TC}-H{r}")
        entrada(pg[f"L{r}"], fmt=FECHA)
        entrada(pg[f"M{r}"], fmt=MONEDA)
        money(pg[f"N{r}"], f"=C{r}+D{r}*{TC}+H{r}+K{r}", bold=True)
        money(pg[f"O{r}"], f'=IF(E{r}="",C{r}+D{r}*{TC},0)+IF(I{r}="",H{r},0)+IF(L{r}="",K{r},0)', color=ROJO)
        pg[f"P{r}"] = (f'=IF(N{r}=0,"—",IF(O{r}=0,"✅ Todo pagado",IF($A{r}<Pagos!$R$1,"⚠️ Atrasado",'
                       f'IF($A{r}=Pagos!$R$1,"⏳ Pagar este mes","📅 Próximo"))))')
        style(pg[f"P{r}"], h="center")
        money(pg[f"Q{r}"], (f'=IF(AND(E{r}<>"",F{r}<>""),F{r}-C{r},0)+IF(AND(E{r}<>"",G{r}<>""),(G{r}-D{r})*{TC},0)'
                            f'+IF(AND(I{r}<>"",J{r}<>""),J{r}-H{r},0)+IF(AND(L{r}<>"",M{r}<>""),M{r}-K{r},0)'), color=GRIS)
        pg.row_dimensions[r].height = 20
        m += 1
        if m == 13:
            y, m = y + 1, 1
    # Marcar como pagados los meses anteriores al actual (la planilla se creó con esas cuotas ya pagas)
    grupos = {"Crédito": "E", "MercadoPago": "I"}
    for r in range(5, 103):
        d, n, f, medio = cuo[f"D{r}"].value, cuo[f"E{r}"].value, cuo[f"F{r}"].value, cuo[f"C{r}"].value
        if not d or not n:
            continue
        col = grupos.get(medio, "L")
        ini = idx(d.year, d.month)
        for k in range(ini, min(ini + int(n) - 1, hoy_idx - 1) + 1):
            row = FIRST + (k - idx(anio - 1, 1))
            if FIRST <= row <= LAST:
                pg[f"{col}{row}"] = "✔"
                pg[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
    for (yy, mm), col, val in PAGOS_REALES:
        row = FIRST + (idx(yy, mm) - idx(anio - 1, 1))
        if FIRST <= row <= LAST:
            pg[f"{col}{row}"] = val
            if isinstance(val, str) and val == "✔":
                pg[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
    tr = LAST + 2
    for rr in (tr, tr + 1):
        pg.merge_cells(f"B{rr}:M{rr}")
        for col in "CDEFGHIJKLM":
            style(pg[f"{col}{rr}"], bg=CELESTE)
    label(pg[f"B{tr}"], "Pendiente de pago (meses vencidos y el actual)", bold=True, bg=CELESTE)
    money(pg[f"N{tr}"], f"=SUMPRODUCT(($A${FIRST}:$A${LAST}<=Pagos!$R$1)*$O${FIRST}:$O${LAST})", bold=True, color=ROJO, bg=CELESTE)
    label(pg[f"B{tr+1}"], "Comprometido en meses futuros", bold=True, bg=CELESTE)
    money(pg[f"N{tr+1}"], f"=SUMPRODUCT(($A${FIRST}:$A${LAST}>Pagos!$R$1)*$N${FIRST}:$N${LAST})", bold=True, bg=CELESTE)
    pg.merge_cells(f"B{tr+3}:Q{tr+3}")
    pg.row_dimensions[tr + 3].height = 30
    note(pg, f"B{tr+3}", "Cubre desde enero del año anterior hasta diciembre del año siguiente al de Configuración. "
         "Si una compra termina después, agregá filas copiando la última. 'Crédito (USD)' es lo que tenés que comprar en "
         "dólares ese mes; en 'USD pagados' anotá lo que compraste de verdad.")
    pg.conditional_formatting.add(f"B{FIRST}:Q{LAST}", FormulaRule(formula=[f"$A{FIRST}=Pagos!$R$1"], fill=fill("FFE8F0FA")))
    pg.conditional_formatting.add(f"P{FIRST}:P{LAST}", FormulaRule(formula=[f'ISNUMBER(SEARCH("Atrasado",P{FIRST}))'], font=Font(color=ROJO, bold=True), fill=fill("FFF8CBAD")))
    pg.conditional_formatting.add(f"P{FIRST}:P{LAST}", FormulaRule(formula=[f'ISNUMBER(SEARCH("Todo pagado",P{FIRST}))'], font=Font(color=VERDE, bold=True)))
    pg.conditional_formatting.add(f"P{FIRST}:P{LAST}", FormulaRule(formula=[f'ISNUMBER(SEARCH("Pagar este mes",P{FIRST}))'], font=Font(color=ROJO, bold=True)))
    pg.conditional_formatting.add(f"Q{FIRST}:Q{LAST}", CellIsRule(operator="notEqual", formula=["0"], font=Font(color="FFB9770E", bold=True)))
    pg.column_dimensions["A"].hidden = True
    for col, w in zip("BCDEFGHIJKLMNOPQ", [17, 15, 13, 13, 15, 11, 15, 13, 15, 14, 13, 15, 17, 14, 18, 15]):
        pg.column_dimensions[col].width = w
    pg.freeze_panes = "C5"

    # ----------------------------------------------------------------- Cuotas
    for desc, col, val in AJUSTES_CUOTAS:
        for r in range(5, 103):
            if str(cuo[f"A{r}"].value or "").strip().lower() == desc.lower():
                cuo[f"{col}{r}"] = val
                break
    for r, col, val in AJUSTES_CUOTAS_POR_FILA:
        cuo[f"{col}{r}"] = val
    # columna R: cuota en dólares (opcional). Si está, F (pesos) se calcula con el dólar de Configuración.
    cuo["R4"] = "Cuota en USD\n(si se paga en dólares)"
    cuo["R4"]._style = copy(cuo["L4"]._style)
    for r in range(5, 103):
        entrada(cuo[f"R{r}"], fmt='"USD "#,##0.00')
    def poner_usd(r, usd):
        cuo[f"R{r}"] = usd
        cuo[f"F{r}"] = f"=R{r}*{TC}"
        style(cuo[f"F{r}"], fmt=MONEDA, color=AZUL_TITULO)
    for desc, usd in CUOTAS_USD.items():
        for r in range(5, 103):
            if str(cuo[f"A{r}"].value or "").strip().lower() == desc.lower():
                poner_usd(r, usd)
                break
    ultima = max(r for r in range(5, 103) if cuo[f"A{r}"].value)
    for desc, cat, medio, fecha, n, pesos, usd in CUOTAS_NUEVAS:
        ultima += 1
        r = ultima
        cuo[f"A{r}"], cuo[f"B{r}"], cuo[f"C{r}"], cuo[f"D{r}"], cuo[f"E{r}"] = desc, cat, medio, fecha, n
        if usd is not None:
            poner_usd(r, usd)
        else:
            cuo[f"F{r}"] = pesos
    # Validaciones heredadas rotas (#REF! y una "custom" sin fórmula): Excel las rechaza.
    from openpyxl.worksheet.datavalidation import DataValidation
    cuo.data_validations.dataValidation = [
        dv for dv in cuo.data_validations.dataValidation
        if not ("B5:C102" in str(dv.sqref) or "F5:G102" in str(dv.sqref))
    ]
    for rng, src in (("B5:B102", "'Configuración'!$A$7:$A$34"), ("C5:C102", "'Configuración'!$E$7:$E$13")):
        dv = DataValidation(type="list", formula1=src, allow_blank=True, showErrorMessage=False)
        dv.add(rng)
        cuo.add_data_validation(dv)
    cuo["A2"] = ("Completá sólo lo amarillo (A–F). El total, el período y el estado se calculan solos y cada cuota "
                 "se suma al Tablero y al Resumen del mes que corresponde. 'Pagadas' y 'Faltan' salen de las marcas "
                 "que hacés en la hoja Pagos cuando pagás cada tarjeta. Lo que pagás en dólares va en 'Cuota en USD' (columna R): "
                 "el valor en pesos se calcula solo con el dólar de Configuración.")
    cuo["K4"] = "Faltan"
    # La "Tabla_1" y el autofiltro que exportó Google Sheets se superponen en A4:L102 y
    # Excel los rechaza al abrir ("Característica quitada: Tabla"). Se dejan como rango común.
    if "Tabla_1" in cuo.tables:
        del cuo.tables["Tabla_1"]
    cuo.auto_filter.ref = None
    cuo["P4"] = "Restante ($)"
    cuo["Q4"] = "Próxima a pagar"
    for c in ("P4", "Q4"):
        cuo[c]._style = copy(cuo["L4"]._style)
    PA = f"Pagos!$A${FIRST}:$A${LAST}"
    for r in range(5, 103):
        pagos_marca = (f'(($C{r}="Crédito")*(Pagos!$E${FIRST}:$E${LAST}<>"")'
                       f'+($C{r}="MercadoPago")*(Pagos!$I${FIRST}:$I${LAST}<>"")'
                       f'+($C{r}<>"Crédito")*($C{r}<>"MercadoPago")*(Pagos!$L${FIRST}:$L${LAST}<>""))')
        cuo[f"J{r}"] = f'=IF($D{r}="","",SUMPRODUCT(({PA}>=$M{r})*({PA}<=$N{r})*{pagos_marca}))'
        cuo[f"K{r}"] = f'=IF($D{r}="","",$E{r}-$J{r})'
        cuo[f"L{r}"] = (f'=IF($D{r}="","",IF($J{r}>=$E{r},"✅ Finalizado",'
                        f'IF(MIN($E{r},Pagos!$R$1-$M{r})>$J{r},"⚠️ Atrasada "&$J{r}&"/"&$E{r},'
                        f'IF($M{r}>Pagos!$R$1,"⏳ Por comenzar","🔵 En curso "&$J{r}&"/"&$E{r}))))')
        cuo[f"P{r}"] = f'=IF($D{r}="","",$K{r}*$F{r})'
        cuo[f"Q{r}"] = f'=IF($D{r}="","",IF($K{r}=0,"—",TEXT(EDATE($D{r},$J{r}),"MM/YYYY")))'
        cuo[f"K{r}"]._style = copy(cuo[f"J{r}"]._style)
        cuo[f"P{r}"]._style = copy(cuo[f"G{r}"]._style)
        cuo[f"Q{r}"]._style = copy(cuo[f"H{r}"]._style)
    cuo.conditional_formatting._cf_rules.clear()
    cuo.conditional_formatting.add("L5:L102", FormulaRule(formula=['ISNUMBER(SEARCH("Finalizado",L5))'], font=Font(color=VERDE, bold=True)))
    cuo.conditional_formatting.add("L5:L102", FormulaRule(formula=['ISNUMBER(SEARCH("Por comenzar",L5))'], font=Font(color=GRIS)))
    cuo.conditional_formatting.add("L5:L102", FormulaRule(formula=['ISNUMBER(SEARCH("Atrasada",L5))'], font=Font(color=ROJO, bold=True), fill=fill("FFF8CBAD")))
    cuo.column_dimensions["L"].width = 20
    cuo.column_dimensions["P"].width = 15
    cuo.column_dimensions["Q"].width = 15
    cuo.column_dimensions["R"].width = 16

    # ---------------------------------------------------------------- Tablero
    tab["C18"] = tab["C17"].value.replace("$A17", "$A18")
    tab["C18"]._style = copy(tab["C17"]._style)
    tab["A37"] = "TOTAL GASTOS"
    tab["C37"] = '=SUMIF($B$10:$B$36,"Gasto",C10:C36)'
    tab["D37"] = '=SUMIF($B$10:$B$36,"Gasto",D10:D36)'
    tab["E37"] = '=IF(N(D37)=0,"",D37-C37)'
    tab["F37"] = '=IF(N(D37)=0,"",C37/D37)'
    for c in ("E37", "F37"):
        tab[c]._style = copy(tab["C37"]._style)
    tab["F37"].number_format = "0.0%"
    header(tab["E5"], "💳 Crédito del mes ($)", bg=AZUL_HEADER)
    header(tab["F5"], "🟡 MercadoPago del mes", bg=AZUL_HEADER)
    header(tab["G5"], "👛 Billetera hoy", bg=AZUL_HEADER)
    money(tab["E6"], f"=SUMIFS(Pagos!$C${FIRST}:$C${LAST},{PA},$H$2)", bold=True, size=14, color=ROJO, bg=CELESTE)
    money(tab["F6"], f"=SUMIFS(Pagos!$H${FIRST}:$H${LAST},{PA},$H$2)", bold=True, size=14, color=ROJO, bg=CELESTE)
    money(tab["G6"], "=Billetera!$B$15", bold=True, size=14, color=AZUL_TITULO, bg=CELESTE)
    for c in "EFG":
        tab[f"{c}6"].alignment = Alignment(horizontal="center", vertical="center")
    tab["E7"] = (f'=IFERROR(IF(E6+SUMIFS(Pagos!$D${FIRST}:$D${LAST},{PA},$H$2)=0,"sin cuotas","+ USD "&TEXT(SUMIFS(Pagos!$D${FIRST}:$D${LAST},{PA},$H$2),"0.00")'
                 f'&IF(INDEX(Pagos!$E${FIRST}:$E${LAST},MATCH($H$2,{PA},0))<>""," · ✅ pagado"," · ⏳ pendiente")),"")')
    tab["F7"] = (f'=IFERROR(IF(F6=0,"sin cuotas",IF(INDEX(Pagos!$I${FIRST}:$I${LAST},MATCH($H$2,{PA},0))<>"",'
                 f'"✅ pagado","⏳ pendiente")),"")')
    tab["G7"] = "→ detalle en hoja Billetera"
    for c in "EFG":
        style(tab[f"{c}7"], size=9, color=GRIS, h="center", border=False, italic=True)
    tab.column_dimensions["E"].width = 18
    tab.column_dimensions["F"].width = 20
    tab.column_dimensions["G"].width = 20

    # -------------------------------------------------------------- Billetera
    bi = wb.create_sheet("Billetera")
    bi.sheet_view.showGridLines = False
    bi.merge_cells("A1:G1")
    title(bi, "A1", "👛  Billetera  —  cuánta plata deberías tener hoy")
    bi.row_dimensions[1].height = 30
    bi.merge_cells("A2:G2")
    note(bi, "A2", "Suma todos tus ingresos, resta los gastos sueltos, el ahorro que apartaste y las tarjetas que marcaste "
         "como pagadas en Pagos, y le suma el fondo de Aston Birra porque esa plata está en tu poder. "
         "Completá las celdas amarillas: el saldo inicial y, cuando quieras controlar, lo que contaste de verdad.")
    bi.row_dimensions[2].height = 48
    bi["H1"] = "=Pagos!$R$1"                       # mes actual
    bi["H2"] = ("=IF(COUNT(Movimientos!$A:$A)=0,$H$1,YEAR(MIN(Movimientos!$A:$A))*12"
                "+MONTH(MIN(Movimientos!$A:$A)))")  # primer mes cargado
    bi["I1"], bi["I2"] = "hoy (auxiliar)", "primer mes (auxiliar)"

    label(bi["A4"], "Saldo inicial: plata que tenías antes del primer movimiento cargado", bold=True)
    entrada(bi["B4"], 0, fmt=MONEDA)
    label(bi["A5"], "Fecha de hoy")
    bi["B5"] = "=TODAY()"
    style(bi["B5"], fmt=FECHA)
    label(bi["A6"], "Primer mes cargado en Movimientos")
    bi["B6"] = '=TEXT(DATE(INT(($H$2-1)/12),MOD($H$2-1,12)+1,1),"MM/YYYY")'
    style(bi["B6"], h="center")

    header(bi["A8"], "💰  Lo que deberías tener hoy", bg=VERDE)
    header(bi["B8"], "Monto", bg=VERDE)
    HOY = '"<="&TODAY()'
    label(bi["A9"], "Saldo inicial", indent=True)
    money(bi["B9"], "=$B$4")
    label(bi["A10"], "+ Ingresos cargados hasta hoy", indent=True)
    money(bi["B10"], f'=SUMIFS(Movimientos!$F:$F,Movimientos!$B:$B,"Ingreso",Movimientos!$A:$A,{HOY})', color=VERDE)
    label(bi["A11"], "− Gastos sueltos hasta hoy (Movimientos)", indent=True)
    money(bi["B11"], f'=SUMIFS(Movimientos!$F:$F,Movimientos!$B:$B,"Gasto",Movimientos!$A:$A,{HOY})', color=ROJO)
    label(bi["A12"], "− Ahorro apartado (no cuenta como disponible)", indent=True)
    money(bi["B12"], f'=SUMIFS(Movimientos!$F:$F,Movimientos!$B:$B,"Ahorro",Movimientos!$A:$A,{HOY})', color=ROJO)
    label(bi["A13"], "− Tarjetas y cuotas ya pagadas (monto real + USD pagados al dólar de Configuración)", indent=True)
    money(bi["B13"], f'=SUMPRODUCT(({PA}>=$H$2)*((Pagos!$E${FIRST}:$E${LAST}<>"")*(Pagos!$F${FIRST}:$F${LAST}+(Pagos!$F${FIRST}:$F${LAST}="")*Pagos!$C${FIRST}:$C${LAST}+(Pagos!$G${FIRST}:$G${LAST}+(Pagos!$G${FIRST}:$G${LAST}="")*Pagos!$D${FIRST}:$D${LAST})*{TC})'
                     f'+(Pagos!$I${FIRST}:$I${LAST}<>"")*(Pagos!$J${FIRST}:$J${LAST}+(Pagos!$J${FIRST}:$J${LAST}="")*Pagos!$H${FIRST}:$H${LAST})'
                     f'+(Pagos!$L${FIRST}:$L${LAST}<>"")*(Pagos!$M${FIRST}:$M${LAST}+(Pagos!$M${FIRST}:$M${LAST}="")*Pagos!$K${FIRST}:$K${LAST})))', color=ROJO)
    label(bi["A14"], "+ Fondo Aston Birra (está en tu poder)", indent=True)
    money(bi["B14"], "='Aston Birra'!$B$6", color=VERDE)
    label(bi["A15"], "👛 TOTAL: deberías tener hoy", bold=True, bg=CELESTE)
    money(bi["B15"], "=B9+B10-B11-B12-B13+B14", bold=True, size=14, color=AZUL_TITULO, bg=CELESTE)
    bi.row_dimensions[15].height = 30
    label(bi["A16"], "de eso, es plata del club Aston Birra", color=GRIS, indent=True)
    money(bi["B16"], "=B14", color=GRIS)
    label(bi["A17"], "plata propia", color=GRIS, indent=True)
    money(bi["B17"], "=B15-B14", color=GRIS)

    header(bi["A19"], "💳  Lo que todavía tenés que pagar", bg=AZUL_HEADER)
    header(bi["B19"], "Monto", bg=AZUL_HEADER)
    label(bi["A20"], "Crédito pendiente (meses vencidos y actual; USD al dólar de Configuración)", indent=True)
    money(bi["B20"], f'=SUMPRODUCT(({PA}<=$H$1)*(Pagos!$E${FIRST}:$E${LAST}="")*(Pagos!$C${FIRST}:$C${LAST}+Pagos!$D${FIRST}:$D${LAST}*{TC}))', color=ROJO)
    label(bi["A21"], "MercadoPago pendiente", indent=True)
    money(bi["B21"], f'=SUMPRODUCT(({PA}<=$H$1)*(Pagos!$I${FIRST}:$I${LAST}="")*Pagos!$H${FIRST}:$H${LAST})', color=ROJO)
    label(bi["A22"], "Préstamo / otros pendientes (hoja Pagos)", indent=True)
    money(bi["B22"], f'=SUMPRODUCT(({PA}<=$H$1)*(Pagos!$L${FIRST}:$L${LAST}="")*Pagos!$K${FIRST}:$K${LAST})', color=ROJO)
    # filas de los bloques de abajo
    RC0 = 27                       # "Por cobrar": encabezado
    RC1, RC2 = RC0 + 1, RC0 + 6    # filas editables
    RP0 = RC2 + 4                  # "Previstos": encabezado
    RP1, RP2 = RP0 + 1, RP0 + 6
    RK0 = RP2 + 3                  # control de caja
    RT0 = RK0 + 7                  # tabla mensual: título
    label(bi["A23"], "Facturas e impuestos previstos del mes (bloque de abajo)", indent=True)
    money(bi["B23"], f"=B{RP2+1}", color=ROJO)
    label(bi["A24"], "✅ DISPONIBLE REAL después de pagar todo", bold=True, bg=CELESTE)
    money(bi["B24"], "=B15-B20-B21-B22-B23", bold=True, size=14, color=VERDE, bg=CELESTE)
    bi.row_dimensions[24].height = 30
    label(bi["A25"], "Comprometido en cuotas de meses futuros (info)", color=GRIS, indent=True)
    money(bi["B25"], f"=SUMPRODUCT(({PA}>$H$1)*Pagos!$N${FIRST}:$N${LAST})", color=GRIS)

    header(bi[f"A{RC0}"], "📥  Por cobrar: plata que te deben y todavía no tenés (cuando la recibas, cargala en Movimientos y borrala de acá)", bg=AZUL_HEADER)
    header(bi[f"B{RC0}"], "Monto", bg=AZUL_HEADER)
    for i, r in enumerate(range(RC1, RC2 + 1)):
        entrada(bi[f"A{r}"]); entrada(bi[f"B{r}"], fmt=MONEDA)
        bi[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
        if i < len(POR_COBRAR):
            bi[f"A{r}"], bi[f"B{r}"] = POR_COBRAR[i]
    label(bi[f"A{RC2+1}"], "Total por cobrar", bold=True)
    money(bi[f"B{RC2+1}"], f"=SUM(B{RC1}:B{RC2})", bold=True, color=VERDE)
    label(bi[f"A{RC2+2}"], "Disponible real si cobrás todo", bold=True, bg=CELESTE)
    money(bi[f"B{RC2+2}"], f"=B24+B{RC2+1}", bold=True, color=VERDE, bg=CELESTE)

    header(bi[f"A{RP0}"], "📤  Facturas e impuestos previstos del mes (estimación editable; cuando pagues, cargalo en Movimientos y borralo de acá)", bg=AZUL_HEADER)
    header(bi[f"B{RP0}"], "Monto", bg=AZUL_HEADER)
    for i, r in enumerate(range(RP1, RP2 + 1)):
        entrada(bi[f"A{r}"]); entrada(bi[f"B{r}"], fmt=MONEDA)
        bi[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
        if i < len(PREVISTOS):
            bi[f"A{r}"], bi[f"B{r}"] = PREVISTOS[i]
    label(bi[f"A{RP2+1}"], "Total previsto", bold=True)
    money(bi[f"B{RP2+1}"], f"=SUM(B{RP1}:B{RP2})", bold=True, color=ROJO)

    header(bi[f"A{RK0}"], "🔎  Control de caja: contá lo que tenés y compará (si da negativo, hay gastos sin cargar)", bg=AZUL_HEADER)
    header(bi[f"B{RK0}"], "Monto", bg=AZUL_HEADER)
    for r, t in ((RK0 + 1, "Efectivo que contaste"), (RK0 + 2, "Banco / débito"), (RK0 + 3, "MercadoPago / billeteras")):
        label(bi[f"A{r}"], t, indent=True)
        entrada(bi[f"B{r}"], fmt=MONEDA)
    label(bi[f"A{RK0+4}"], "Total real contado", bold=True)
    money(bi[f"B{RK0+4}"], f"=SUM(B{RK0+1}:B{RK0+3})", bold=True)
    label(bi[f"A{RK0+5}"], "Diferencia real − teórico", bold=True, bg=CELESTE)
    money(bi[f"B{RK0+5}"], f'=IF(B{RK0+4}=0,"",B{RK0+4}-B15)', bold=True, bg=CELESTE)
    bi.conditional_formatting.add(f"B{RK0+5}", CellIsRule(operator="lessThan", formula=["0"], font=Font(color=ROJO, bold=True)))

    header(bi[f"A{RT0}"], "📅  Sobrante mes a mes (año de Configuración)", bg=VERDE)
    bi.merge_cells(f"A{RT0}:G{RT0}")
    for i, h in enumerate(["Mes", "Ingresos", "Gastos sueltos", "Cuotas del mes", "Ahorro",
                           "Sobrante del mes", "Acumulado\n(plata propia)"], 1):
        header(bi.cell(RT0 + 1, i), h)
    bi.row_dimensions[RT0 + 1].height = 30
    M1 = RT0 + 2
    for i, nombre in enumerate(MESES, 1):
        r = M1 + i - 1
        bi[f"H{r}"] = f"='Configuración'!$B$3*12+{i}"
        bi[f"I{r}"] = f'=TEXT(DATE(\'Configuración\'!$B$3,{i},1),"YYYY-MM")'
        label(bi[f"A{r}"], nombre, bold=True)
        cond = f'IF(OR($H{r}<$H$2,$H{r}>$H$1),"",'
        money(bi[f"B{r}"], f'={cond}SUMIFS(Movimientos!$F:$F,Movimientos!$G:$G,$I{r},Movimientos!$B:$B,"Ingreso"))')
        money(bi[f"C{r}"], f'={cond}SUMIFS(Movimientos!$F:$F,Movimientos!$G:$G,$I{r},Movimientos!$B:$B,"Gasto"))')
        money(bi[f"D{r}"], f'={cond}SUMIFS(Pagos!$N${FIRST}:$N${LAST},{PA},$H{r}))')
        money(bi[f"E{r}"], f'={cond}SUMIFS(Movimientos!$F:$F,Movimientos!$G:$G,$I{r},Movimientos!$B:$B,"Ahorro"))')
        money(bi[f"F{r}"], f'=IF(B{r}="","",B{r}-C{r}-D{r}-E{r})', bold=True)
        money(bi[f"G{r}"], f'=IF(F{r}="","",$B$4+SUM($F${M1}:F{r}))', bold=True, color=AZUL_TITULO)
    M12 = M1 + 11
    label(bi[f"A{M12+1}"], "TOTAL AÑO", bold=True, bg=CELESTE)
    for c in "BCDEF":
        money(bi[f"{c}{M12+1}"], f"=SUM({c}{M1}:{c}{M12})", bold=True, bg=CELESTE)
    style(bi[f"G{M12+1}"], bg=CELESTE)
    note(bi, f"A{M12+2}", "Sólo se muestran los meses desde que empezaste a cargar hasta el mes actual. "
         "'Acumulado' = saldo inicial + sobrantes de cada mes: es tu plata propia si pagás todas las cuotas del mes.")
    bi.merge_cells(f"A{M12+2}:G{M12+2}")
    bi.row_dimensions[M12 + 2].height = 36
    bi.conditional_formatting.add(f"F{M1}:G{M12}", CellIsRule(operator="lessThan", formula=["0"], font=Font(color=ROJO, bold=True)))
    bi.column_dimensions["A"].width = 62
    for c in "BCDEFG":
        bi.column_dimensions[c].width = 17
    bi.column_dimensions["H"].hidden = True
    bi.column_dimensions["I"].hidden = True
    bi.freeze_panes = "A4"

    # ------------------------------------------------------------ Aston Birra
    ast["A2"] = ("Hoja independiente: los aportes NO afectan tu Tablero personal, pero el saldo del fondo sí se suma "
                 "en tu Billetera porque la plata está en tu poder. Cargá cada aporte a la derecha; el saldo, la meta "
                 "y el total por persona se calculan solos.")

    # ---------------------------------------------------------- Instrucciones
    r0 = ins.max_row + 2
    ins[f"A{r0}"] = "Novedades: Billetera y Pagos"
    ins[f"A{r0}"]._style = copy(ins["A3"]._style)
    ins.merge_cells(f"A{r0}:B{r0}")
    ins.row_dimensions[r0].height = 24
    lineas = [
        "👛 Billetera → te dice cuánta plata deberías tener hoy: saldo inicial + ingresos − gastos − ahorro − tarjetas pagadas + fondo Aston Birra.",
        "   Cargá una vez el 'Saldo inicial' (lo que tenías antes del primer movimiento). Abajo ves el sobrante de cada mes y cómo se acumula.",
        "   'Control de caja': contá efectivo + banco + MercadoPago y compará con lo teórico. Si da negativo, hay gastos que no cargaste.",
        "💳 Pagos → una fila por mes con lo que debés pagar de Crédito, MercadoPago y Otros (sale solo de Cuotas).",
        "   Cuando pagás el resumen, escribí la fecha (o ✔) en 'Pagado el' y el importe real en 'Monto real' (si fue distinto a lo calculado: impuestos, comisiones, dólares).",
        "   'Diferencia' te muestra cuánto pagaste de más o de menos respecto de lo cargado en Cuotas: es lo que te falta cargar.",
        "💵 Dólares: lo que pagás en USD va en Cuotas columna 'Cuota en USD'. Pagos te dice cuántos USD comprar cada mes y en Configuración ponés el dólar del día.",
        "💳 Cuotas → 'Pagadas', 'Faltan', 'Restante ($)' y 'Próxima a pagar' ahora salen de las marcas de Pagos. '⚠️ Atrasada' = hay un mes viejo sin marcar.",
        "📊 Tablero → arriba a la derecha ves Crédito y MercadoPago del mes elegido (con ✅/⏳ según lo hayas pagado) y tu Billetera de hoy.",
        "Circuito sugerido: 1) cargás movimientos a diario · 2) cuando cae el resumen, marcás el pago en Pagos · 3) miras Billetera para saber cuánto te queda.",
    ]
    for i, t in enumerate(lineas, 1):
        ins[f"A{r0+i}"] = t
        ins[f"A{r0+i}"]._style = copy(ins["A4"]._style)
        ins.row_dimensions[r0 + i].height = 19.5

    # ---------------------------------------------------------- orden de hojas
    orden = ["Movimientos", "Cuotas", "Pagos", "Tablero", "Billetera", "Resumen Anual",
             "Aston Birra", "Configuración", "Instrucciones"]
    wb._sheets = [wb[n] for n in orden]
    wb.active = orden.index("Tablero")
    wb.calculation.fullCalcOnLoad = True
    wb.save(dst)
    print("Guardado:", dst)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    main(sys.argv[1], sys.argv[2])
